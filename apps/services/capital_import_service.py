"""
Upload / import shareholder capital register from CSV or Excel.

Accepts English headers and common Somali Excel headers used by Akram Sweets:
  Magaca, Saamiga, Lacagta, Boqoleey %
"""

from __future__ import annotations

import csv
import io
import re
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from typing import BinaryIO, Optional

from apps import db
from apps.models.shareholder import OwnershipRecord, Shareholder
from apps.services.share_value_service import save_share_settings

MONEY = Decimal('0.01')
SHARES = Decimal('0.0001')
OWNERSHIP = Decimal('0.0001')
OWNERSHIP_TOLERANCE = Decimal('0.05')

# Normalized header → field
HEADER_ALIASES = {
    # name
    'name': 'name',
    'magaca': 'name',
    'shareholder': 'name',
    'shareholder name': 'name',
    'full name': 'name',
    # email
    'email': 'email',
    'e-mail': 'email',
    'mail': 'email',
    # shares
    'shares': 'shares',
    'share': 'shares',
    'share_count': 'shares',
    'saamiga': 'shares',
    'saami': 'shares',
    # capital / assets
    'capital': 'capital',
    'investment': 'capital',
    'investment_amount': 'capital',
    'assets': 'capital',
    'lacagta': 'capital',
    'lacagta capital': 'capital',
    'amount': 'capital',
    # ownership
    'ownership_percent': 'ownership_percent',
    'ownership': 'ownership_percent',
    'ownership %': 'ownership_percent',
    'percent': 'ownership_percent',
    'boqoleey': 'ownership_percent',
    'boqoleey %': 'ownership_percent',
    'boqoley': 'ownership_percent',
    'boqoley %': 'ownership_percent',
    # optional
    'is_owner': 'is_owner',
    'owner': 'is_owner',
    'phone': 'phone',
    'mobile': 'phone',
    'telephone': 'phone',
    'country': 'country',
    'country_code': 'country_code',
}

TEMPLATE_HEADERS = (
    'name',
    'email',
    'shares',
    'capital',
    'ownership_percent',
    'is_owner',
    'phone',
    'country',
    'country_code',
)

# Excel footer / note rows that must never become shareholders
_SKIP_NAME_RE = re.compile(
    r'('
    r'\btotal\b|'
    r'\bwadarta\b|'
    r'\bsum\b|'
    r'grand\s*total|'
    r'total\s*asset|'
    r'intaas\s+ayay|'
    r'noqonaysaa|'
    r'faaiidada\s+shirkad|'
    r'dadka\s+qaar|'
    r'faahfaahin|'
    r'boqoleyda\s+lagaf'
    r')',
    re.IGNORECASE,
)

# Company Murabaha / company-owned asset note rows → setting, not a shareholder
_MURABAHA_NAME_RE = re.compile(
    r'('
    r'muraabax|'
    r'murabaha|'
    r'murabaax|'
    r'assetka\s+shirkad|'
    r'company[- ]owned|'
    r'company\s+asset'
    r')',
    re.IGNORECASE,
)


def _normalize_header(raw: str) -> str:
    text = (raw or '').strip().lower()
    text = text.replace('_', ' ')
    text = re.sub(r'\s+', ' ', text)
    # strip currency noise from header cells
    text = text.replace('(usd)', '').replace('$', '').strip()
    return text


def _map_header(raw: str) -> Optional[str]:
    key = _normalize_header(raw)
    if key in HEADER_ALIASES:
        return HEADER_ALIASES[key]
    # "Lacagta" appears twice in Excel (capital then profit) — first mapped wins in column scan
    if key.startswith('lacagta'):
        return 'capital'
    if 'boqol' in key:
        return 'ownership_percent'
    if 'saami' in key:
        return 'shares'
    if key in ('magaca',) or 'name' in key:
        return 'name'
    return None


def _parse_decimal(raw, default=None) -> Optional[Decimal]:
    if raw is None:
        return default
    if isinstance(raw, Decimal):
        return raw
    if isinstance(raw, (int, float)):
        return Decimal(str(raw))
    text = str(raw).strip()
    if not text or text in ('-', '—', 'N/A', 'n/a'):
        return default
    # $ 571,505.00 → 571505.00
    text = text.replace('$', '').replace(',', '').replace('%', '').strip()
    text = re.sub(r'\s+', '', text)
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return default


def _parse_bool(raw) -> bool:
    if isinstance(raw, bool):
        return raw
    return str(raw or '').strip().lower() in ('1', 'true', 'yes', 'y', 'owner')


def _classify_name(name: str) -> str:
    """
    Return 'shareholder', 'murabaha_asset', or 'skip'.

    Akram Excel ends with note rows such as:
      - muraabaxa waaye oo ku jirta assetka shirkada  ($423,000)
      - total asset intaas ayay noqonaysaa             ($1,643,000)
    """
    text = (name or '').strip()
    if not text:
        return 'skip'
    if _MURABAHA_NAME_RE.search(text):
        return 'murabaha_asset'
    if _SKIP_NAME_RE.search(text):
        return 'skip'
    # Pure numeric / punctuation labels
    if re.fullmatch(r'[\d\W_]+', text):
        return 'skip'
    return 'shareholder'


def _slug_email(name: str, used: set) -> str:
    base = re.sub(r'[^a-z0-9]+', '.', (name or 'shareholder').lower()).strip('.')
    if not base:
        base = 'shareholder'
    candidate = f'{base}@akramsweets.com'
    n = 2
    while candidate in used:
        candidate = f'{base}{n}@akramsweets.com'
        n += 1
    used.add(candidate)
    return candidate


def build_template_csv() -> str:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(TEMPLATE_HEADERS)
    writer.writerow([
        'Abdwahaab Said Andirahman',
        'abdwahaab@akramsweets.com',
        '571',
        '571505.00',
        '46.8447',
        'true',
        '',
        'Somalia',
        'so',
    ])
    writer.writerow([
        'Abdiraxman Shikhdoon',
        'abdiraxman.shikhdoon@akramsweets.com',
        '100',
        '100000.00',
        '8.1967',
        'false',
        '',
        'Somalia',
        'so',
    ])
    return buf.getvalue()


def _rows_from_csv(text: str) -> list[dict]:
    # Skip Somali summary lines before the real header table
    lines = text.splitlines()
    header_idx = None
    for i, line in enumerate(lines):
        lower = line.lower()
        if 'magaca' in lower or ('name' in lower and ('saami' in lower or 'share' in lower or 'capital' in lower)):
            header_idx = i
            break
        # English template
        if lower.startswith('name,') or lower.startswith('name;'):
            header_idx = i
            break
    if header_idx is None:
        header_idx = 0
    payload = '\n'.join(lines[header_idx:])
    sample = payload[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=',;\t')
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(payload), dialect)
    rows_raw = list(reader)
    if not rows_raw:
        return []
    headers = rows_raw[0]
    mapped = []
    capital_seen = False
    for h in headers:
        field = _map_header(h)
        if field == 'capital':
            if capital_seen:
                # Second Lacagta column = profit — ignore
                mapped.append(None)
            else:
                capital_seen = True
                mapped.append('capital')
        else:
            mapped.append(field)

    out = []
    for cells in rows_raw[1:]:
        if not any(str(c or '').strip() for c in cells):
            continue
        row = {}
        for idx, field in enumerate(mapped):
            if not field or idx >= len(cells):
                continue
            # Prefer first capital column only
            if field in row and field == 'capital':
                continue
            row[field] = cells[idx]
        name = str(row.get('name') or '').strip()
        if not name:
            continue
        out.append(row)
    return out


def _rows_from_xlsx(data: bytes) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError(
            'Excel support requires openpyxl. Run: pip install openpyxl'
        ) from exc

    wb = load_workbook(filename=io.BytesIO(data), data_only=True, read_only=True)
    ws = wb.active
    grid = []
    for row in ws.iter_rows(values_only=True):
        grid.append(list(row))
    wb.close()
    if not grid:
        return []

    header_idx = 0
    for i, row in enumerate(grid):
        joined = ' '.join(str(c or '') for c in row).lower()
        if 'magaca' in joined or ('name' in joined and ('saami' in joined or 'share' in joined)):
            header_idx = i
            break

    headers = [str(c or '') for c in grid[header_idx]]
    mapped = []
    capital_seen = False
    for h in headers:
        field = _map_header(h)
        if field == 'capital':
            if capital_seen:
                mapped.append(None)
            else:
                capital_seen = True
                mapped.append('capital')
        else:
            mapped.append(field)

    out = []
    for cells in grid[header_idx + 1 :]:
        if not any(c is not None and str(c).strip() for c in cells):
            continue
        row = {}
        for idx, field in enumerate(mapped):
            if not field or idx >= len(cells):
                continue
            row[field] = cells[idx]
        name = str(row.get('name') or '').strip()
        if not name:
            continue
        out.append(row)
    return out


def parse_upload(file_storage) -> list[dict]:
    """Parse Werkzeug FileStorage (CSV or XLSX) into raw row dicts."""
    if not file_storage or not file_storage.filename:
        raise ValueError('Choose a CSV or Excel (.xlsx) file to upload.')

    filename = file_storage.filename.lower().strip()
    data = file_storage.read()
    if not data:
        raise ValueError('Uploaded file is empty.')

    if filename.endswith(('.xlsx', '.xlsm')):
        return _rows_from_xlsx(data)
    if filename.endswith('.xls'):
        raise ValueError('Legacy .xls is not supported. Save as .xlsx or CSV and upload again.')
    # CSV / TSV / text
    for encoding in ('utf-8-sig', 'utf-8', 'latin-1'):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            text = None
    if text is None:
        raise ValueError('Could not read the file encoding. Save as UTF-8 CSV or .xlsx.')
    return _rows_from_csv(text)


def normalize_rows(raw_rows: list[dict]):
    """Validate/normalize rows; derive ownership from capital when missing.

    Returns (rows, warnings, meta).
    meta includes detected company_owned_assets from Murabaha note rows.
    """
    warnings = []
    used_emails: set[str] = set()
    normalized = []
    detected_company_assets = None
    skipped = 0

    for i, raw in enumerate(raw_rows, start=1):
        name = str(raw.get('name') or '').strip()
        if not name:
            warnings.append(f'Row {i}: skipped (missing name).')
            skipped += 1
            continue

        kind = _classify_name(name)
        capital = _parse_decimal(raw.get('capital'), Decimal('0')) or Decimal('0')
        shares = _parse_decimal(raw.get('shares'), Decimal('0')) or Decimal('0')

        if kind == 'murabaha_asset':
            if capital > 0:
                detected_company_assets = capital.quantize(MONEY)
                warnings.append(
                    f'Skipped “{name}” — treated as company-owned Murabaha assets '
                    f'(${detected_company_assets:,.2f}), not a shareholder.'
                )
            else:
                warnings.append(f'Skipped “{name}” (Murabaha / company-asset note).')
            skipped += 1
            continue

        if kind == 'skip':
            warnings.append(f'Skipped “{name}” (summary / total / note row).')
            skipped += 1
            continue

        # Extra guard: huge capital with almost no shares = balance-sheet total line
        if capital >= Decimal('1000000') and shares <= Decimal('1'):
            warnings.append(
                f'Skipped “{name}” — looks like a total-assets line (${capital:,.2f}).'
            )
            skipped += 1
            continue

        ownership = _parse_decimal(raw.get('ownership_percent'), None)

        email = str(raw.get('email') or '').strip().lower()
        if email and '@' in email:
            if email in used_emails:
                email = _slug_email(name, used_emails)
                warnings.append(f'Row {i} ({name}): duplicate email — generated {email}.')
            else:
                used_emails.add(email)
        else:
            email = _slug_email(name, used_emails)
            warnings.append(f'Row {i} ({name}): no email — using {email}.')

        normalized.append({
            'name': name,
            'email': email,
            'shares': shares.quantize(SHARES),
            'capital': capital.quantize(MONEY),
            'ownership_percent': ownership,
            'is_owner': _parse_bool(raw.get('is_owner')),
            'phone': (str(raw.get('phone')).strip() if raw.get('phone') not in (None, '') else None),
            'country': (str(raw.get('country')).strip() if raw.get('country') not in (None, '') else 'Somalia'),
            'country_code': (
                str(raw.get('country_code') or 'so').strip().lower()[:8] or 'so'
            ),
        })

    if not normalized:
        raise ValueError(
            'No shareholder rows found. Remove summary lines and ensure Magaca/Name column exists.'
        )

    # Excel stores 47% as 0.47 — convert fractions to percent when sum ≈ 1
    own_values = [r['ownership_percent'] for r in normalized if r['ownership_percent'] is not None]
    if own_values:
        own_sum = sum(own_values, Decimal('0'))
        if Decimal('0.85') <= own_sum <= Decimal('1.15'):
            for r in normalized:
                if r['ownership_percent'] is not None:
                    r['ownership_percent'] = (r['ownership_percent'] * Decimal('100')).quantize(
                        OWNERSHIP, rounding=ROUND_HALF_UP
                    )
            warnings.append(
                'Excel ownership cells were fractions (e.g. 0.47 = 47%). Converted to percent.'
            )

    total_capital = sum((r['capital'] for r in normalized), Decimal('0'))
    total_shares = sum((r['shares'] for r in normalized), Decimal('0'))

    # Prefer capital-based ownership for the Akram register (precise dollars, rounded Excel %)
    if total_capital > 0:
        for r in normalized:
            r['ownership_percent'] = (r['capital'] / total_capital * Decimal('100')).quantize(
                OWNERSHIP, rounding=ROUND_HALF_UP
            )
        warnings.append(
            'Ownership % calculated from each shareholder’s capital ÷ total shareholder capital '
            '(Murabaha / company assets excluded).'
        )

    for r in normalized:
        if r['ownership_percent'] is None:
            r['ownership_percent'] = Decimal('0')
        r['ownership_percent'] = Decimal(r['ownership_percent']).quantize(
            OWNERSHIP, rounding=ROUND_HALF_UP
        )

    # Rebalance tiny rounding drift onto the largest capital holder
    total_own = sum((r['ownership_percent'] for r in normalized), Decimal('0'))
    drift = (Decimal('100') - total_own).quantize(OWNERSHIP, rounding=ROUND_HALF_UP)
    if abs(drift) > 0 and abs(drift) <= Decimal('0.05'):
        top = max(normalized, key=lambda r: r['capital'])
        top['ownership_percent'] = (top['ownership_percent'] + drift).quantize(OWNERSHIP)
        warnings.append(f'Adjusted {top["name"]} ownership by {drift}% so total is 100%.')
        total_own = Decimal('100')

    if abs(total_own - Decimal('100')) > OWNERSHIP_TOLERANCE:
        warnings.append(
            f'Ownership totals {total_own}% (expected 100%). Fix the file before approving periods.'
        )

    # Mark largest capital as owner if none flagged
    if not any(r['is_owner'] for r in normalized) and normalized:
        top = max(normalized, key=lambda r: r['capital'])
        top['is_owner'] = True
        warnings.append(f'Marked {top["name"]} as owner (largest capital).')

    # Shares from capital when missing (1 share = 1000)
    for r in normalized:
        if r['shares'] <= 0 and r['capital'] > 0:
            r['shares'] = (r['capital'] / Decimal('1000')).quantize(SHARES, rounding=ROUND_HALF_UP)

    # Recompute share total after backfill
    total_shares = sum((r['shares'] for r in normalized), Decimal('0'))

    meta = {
        'row_count': len(normalized),
        'skipped_rows': skipped,
        'total_shares': total_shares,
        'total_capital': total_capital,
        'total_ownership': sum((r['ownership_percent'] for r in normalized), Decimal('0')),
        'company_owned_assets': detected_company_assets,
    }
    return normalized, warnings, meta


def preview_import(file_storage) -> dict:
    raw = parse_upload(file_storage)
    rows, warnings, meta = normalize_rows(raw)
    clean = [{k: v for k, v in r.items() if k != '_meta'} for r in rows]
    return {
        'ok': True,
        'rows': clean,
        'warnings': warnings,
        'meta': meta,
    }


def _norm_name(name: str) -> str:
    return re.sub(r'[^a-z0-9]+', ' ', (name or '').lower()).strip()


def _find_existing_shareholder(row: dict, claimed_ids: set[int]):
    """Match by email first, then by exact normalized name (for re-uploads)."""
    email = (row.get('email') or '').strip().lower()
    if email:
        sh = Shareholder.query.filter_by(email=email).first()
        if sh and sh.id not in claimed_ids:
            return sh

    target = _norm_name(row.get('name') or '')
    if not target:
        return None
    for sh in Shareholder.query.all():
        if sh.id in claimed_ids:
            continue
        if _norm_name(sh.name) == target:
            return sh
    return None


def _close_open_ownership(shareholder_id: int, effective: date):
    open_recs = OwnershipRecord.query.filter_by(
        shareholder_id=shareholder_id,
        effective_to=None,
    ).all()
    for rec in open_recs:
        if rec.effective_from >= effective:
            db.session.delete(rec)
        else:
            rec.effective_to = effective - timedelta(days=1)


def _deactivate_non_imported(keep_ids: set[int], effective: date) -> int:
    """Deactivate every shareholder not in the uploaded register and close ownership."""
    deactivated = 0
    for sh in Shareholder.query.filter_by(is_active=True).all():
        if sh.id in keep_ids and _classify_name(sh.name or '') == 'shareholder':
            continue
        sh.is_active = False
        sh.is_owner = False
        # Zero capital/shares so dashboard totals only reflect the uploaded register
        sh.investment_amount = Decimal('0')
        sh.share_count = Decimal('0')
        _close_open_ownership(sh.id, effective)
        deactivated += 1

    # Also deactivate junk / Murabaha note people left from older imports (even if inactive)
    for sh in Shareholder.query.all():
        if sh.id in keep_ids:
            continue
        if _classify_name(sh.name or '') != 'shareholder' and sh.is_active:
            sh.is_active = False
            sh.is_owner = False
            sh.investment_amount = Decimal('0')
            sh.share_count = Decimal('0')
            _close_open_ownership(sh.id, effective)
            deactivated += 1
    return deactivated


def apply_import(
    rows: list[dict],
    *,
    effective_from: Optional[date] = None,
    update_register_settings: bool = True,
    company_owned_assets: Optional[Decimal] = None,
    actor=None,
) -> dict:
    """
    Authoritative replace of the shareholder capital register.

    - Creates or updates every row in the file (name, shares, capital, ownership)
    - Deactivates anyone not in the file (demo data, old rows, mistaken asset lines)
    - Overwrites company share totals + Murabaha assets settings
    """
    if not rows:
        raise ValueError('No rows to import.')

    effective = effective_from or date.today().replace(day=1)
    created = updated = 0
    keep_ids: set[int] = set()
    claimed_ids: set[int] = set()

    # Clear owner flags before re-applying from file
    for sh in Shareholder.query.filter_by(is_owner=True).all():
        sh.is_owner = False

    for row in rows:
        sh = _find_existing_shareholder(row, claimed_ids)
        if not sh:
            sh = Shareholder(email=row['email'])
            db.session.add(sh)
            created += 1
        else:
            updated += 1
            # Keep a stable login email if one already exists; otherwise use file email
            if not sh.email:
                sh.email = row['email']

        sh.name = row['name']
        # File is authoritative for register contact + capital fields
        if sh.email != row['email']:
            clash = Shareholder.query.filter(
                Shareholder.email == row['email'],
                Shareholder.id != sh.id,
            ).first()
            if not clash:
                sh.email = row['email']
        sh.share_count = row['shares']
        sh.investment_amount = row['capital']
        sh.is_owner = bool(row.get('is_owner'))
        sh.is_active = True
        sh.phone = row.get('phone')
        sh.country = row.get('country') or 'Somalia'
        sh.country_code = row.get('country_code') or 'so'
        db.session.flush()

        claimed_ids.add(sh.id)
        keep_ids.add(sh.id)

        _close_open_ownership(sh.id, effective)
        db.session.add(
            OwnershipRecord(
                shareholder_id=sh.id,
                ownership_percent=row['ownership_percent'],
                effective_from=effective,
                created_by_id=getattr(actor, 'id', None),
            )
        )

    deactivated = _deactivate_non_imported(keep_ids, effective)

    total_shares = sum((Decimal(r['shares']) for r in rows), Decimal('0'))
    total_capital = sum((Decimal(r['capital']) for r in rows), Decimal('0'))

    if update_register_settings:
        from apps.services.share_value_service import get_company_owned_assets, get_share_value

        assets = (
            company_owned_assets
            if company_owned_assets is not None
            else get_company_owned_assets()
        )
        save_share_settings(
            share_value=get_share_value() or Decimal('1000'),
            total_company_shares=total_shares if total_shares > 0 else None,
            company_owned_assets=assets,
        )

    db.session.commit()

    from apps.services.audit_service import log_action

    log_action(
        'import',
        'shareholder_capital',
        None,
        f'Capital register REPLACE: {created} created, {updated} updated, '
        f'{deactivated} removed from active register, {len(rows)} rows, capital={total_capital}',
        user=actor,
    )

    return {
        'ok': True,
        'created': created,
        'updated': updated,
        'deactivated': deactivated,
        'total_rows': len(rows),
        'total_shares': total_shares,
        'total_capital': total_capital,
        'company_owned_assets': company_owned_assets,
        'replaced': True,
    }


def import_from_upload(
    file_storage,
    *,
    effective_from: Optional[date] = None,
    company_owned_assets: Optional[Decimal] = None,
    actor=None,
) -> dict:
    """Parse + fully replace the capital register from an uploaded file."""
    raw = parse_upload(file_storage)
    rows, warnings, meta = normalize_rows(raw)
    clean = [{k: v for k, v in r.items() if k != '_meta'} for r in rows]

    assets = company_owned_assets
    if assets is None and meta.get('company_owned_assets') is not None:
        assets = meta['company_owned_assets']
        warnings.append(
            f'Company-owned assets set from Excel Murabaha row: ${assets:,.2f}.'
        )

    result = apply_import(
        clean,
        effective_from=effective_from,
        update_register_settings=True,
        company_owned_assets=assets,
        actor=actor,
    )
    result['warnings'] = warnings
    result['meta'] = meta
    result['rows'] = clean
    return result

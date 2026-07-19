"""Export monthly profit distribution and capital register to Excel / CSV."""

from __future__ import annotations

import csv
import io
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def _money(value) -> Decimal:
    if value is None:
        return Decimal('0')
    return Decimal(str(value)).quantize(Decimal('0.01'))


def _pct(value) -> Decimal:
    if value is None:
        return Decimal('0')
    return Decimal(str(value)).quantize(Decimal('0.0001'))


def _safe_filename(label: str) -> str:
    cleaned = ''.join(ch if ch.isalnum() or ch in '-_' else '_' for ch in (label or 'export'))
    return cleaned.strip('_') or 'export'


def build_period_distribution_rows(period) -> dict[str, Any]:
    """Collect summary + shareholder rows for a calculated period."""
    from apps.models.period import ShareholderCalculation
    from apps.models.shareholder import Shareholder
    from apps.services.brand_service import get_brand_settings
    from apps.services.mudarabah_service import get_mudarabah_settings
    from apps.services.share_value_service import get_share_settings

    calculations = (
        ShareholderCalculation.query.filter_by(period_id=period.id)
        .join(Shareholder)
        .order_by(Shareholder.name)
        .all()
    )
    if not calculations:
        raise ValueError('Calculate the period before exporting the distribution.')

    brand = get_brand_settings()
    mudarabah = get_mudarabah_settings()
    share_settings = get_share_settings()

    net = _money(period.total_profit_loss)
    pool = _money(period.shareholders_pool)
    partner = _money(period.managing_partner_share)
    alloc_pct = _pct(
        period.mudarabah_shareholder_percent
        if period.mudarabah_shareholder_percent is not None
        else mudarabah.get('shareholder_percent')
    )

    rows = []
    total_shares = Decimal('0')
    total_capital = Decimal('0')
    total_ownership = Decimal('0')
    total_profit = Decimal('0')

    for index, calc in enumerate(calculations, start=1):
        sh = calc.shareholder
        shares = Decimal(str(sh.share_count or 0)) if sh else Decimal('0')
        capital = Decimal(str(sh.investment_amount or 0)) if sh else Decimal('0')
        ownership = _pct(calc.ownership_percent)
        base = _money(calc.base_share)
        arrangement = _money(calc.arrangement_deduction) + _money(calc.arrangement_received)
        manual = _money(calc.manual_adjustment)
        final = _money(calc.final_amount)

        total_shares += shares
        total_capital += capital
        total_ownership += ownership
        total_profit += final

        rows.append({
            'no': index,
            'name': sh.name if sh else '—',
            'shares': shares,
            'capital': capital,
            'ownership_percent': ownership,
            'base_share': base,
            'arrangement_net': arrangement,
            'manual_adjustment': manual,
            'final_amount': final,
        })

    return {
        'company_name': brand.get('company_name') or 'Akram Sweets',
        'period_label': period.period_label,
        'year': period.year,
        'month': period.month,
        'status': period.workflow_label,
        'odoo_reference': period.odoo_reference or '',
        'net_profit': net,
        'allocation_percent': alloc_pct,
        'profit_pool': pool,
        'partner_share': partner,
        'partner_name': mudarabah.get('partner_name') or 'Managing partner',
        'company_owned_assets': _money(share_settings.get('company_owned_assets')),
        'rows': rows,
        'totals': {
            'shares': total_shares,
            'capital': total_capital,
            'ownership_percent': total_ownership,
            'final_amount': total_profit,
            'base_share': sum((r['base_share'] for r in rows), Decimal('0')),
        },
        'shareholder_count': len(rows),
        'filename_base': _safe_filename(f'profit_distribution_{period.period_label}'),
    }


def build_period_distribution_xlsx(period) -> tuple[bytes, str]:
    """Excel workbook shaped like the Shareholders Profit Distribution sheet."""
    data = build_period_distribution_rows(period)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Profit Distribution'

    header_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=11)
    money_font = Font(bold=True)
    thin = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )
    fill_header = PatternFill('solid', fgColor='8A1B24')
    fill_header_font = Font(bold=True, color='FFFFFF')
    fill_summary = PatternFill('solid', fgColor='F8F1F2')
    fill_total = PatternFill('solid', fgColor='EEF2F7')

    ws['A1'] = f'{data["company_name"]} — Shareholders Profit Distribution'
    ws['A1'].font = header_font
    ws.merge_cells('A1:I1')

    ws['A2'] = f'Period: {data["period_label"]}'
    ws['A3'] = f'Status: {data["status"]}'
    if data['odoo_reference']:
        ws['A4'] = f'Odoo reference: {data["odoo_reference"]}'

    ws['A6'] = 'Summary'
    ws['A6'].font = section_font
    ws['A7'] = 'Item'
    ws['B7'] = 'Amount'
    ws['A7'].font = fill_header_font
    ws['B7'].font = fill_header_font
    ws['A7'].fill = fill_header
    ws['B7'].fill = fill_header

    summary_rows = [
        ('Company Net Profit', data['net_profit']),
        (
            f'Profit allocated to shareholders ({data["allocation_percent"]}%)',
            data['profit_pool'],
        ),
        (f'{data["partner_name"]} share', data['partner_share']),
        ('Total distributed (final)', data['totals']['final_amount']),
        ('Active shareholders', data['shareholder_count']),
        ('Total shares', data['totals']['shares']),
        ('Total capital', data['totals']['capital']),
    ]
    for offset, (label, value) in enumerate(summary_rows):
        row_idx = 8 + offset
        ws.cell(row_idx, 1, label).fill = fill_summary
        cell = ws.cell(row_idx, 2, float(value) if not isinstance(value, int) else value)
        cell.fill = fill_summary
        cell.font = money_font
        if isinstance(value, Decimal):
            cell.number_format = '#,##0.00'

    start = 8 + len(summary_rows) + 2
    ws.cell(start, 1, 'Shareholders Profit Distribution').font = section_font

    headers = [
        'No.',
        'Shareholder Name',
        'Shares',
        'Capital Investment',
        'Ownership %',
        'Original Profit (base)',
        'Arrangement net',
        'Manual adjustment',
        'Profit Allocation (final)',
    ]
    header_row = start + 1
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(header_row, col, title)
        cell.font = fill_header_font
        cell.fill = fill_header
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        cell.border = thin

    for i, row in enumerate(data['rows']):
        r = header_row + 1 + i
        values = [
            row['no'],
            row['name'],
            float(row['shares']),
            float(row['capital']),
            float(row['ownership_percent']),
            float(row['base_share']),
            float(row['arrangement_net']),
            float(row['manual_adjustment']),
            float(row['final_amount']),
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(r, col, value)
            cell.border = thin
            if col in (3, 4, 6, 7, 8, 9):
                cell.number_format = '#,##0.00'
            if col == 5:
                cell.number_format = '0.0000'
            if col == 9:
                cell.font = money_font

    total_row = header_row + 1 + len(data['rows'])
    totals = [
        '',
        'TOTAL',
        float(data['totals']['shares']),
        float(data['totals']['capital']),
        float(data['totals']['ownership_percent']),
        float(data['totals']['base_share']),
        '',
        '',
        float(data['totals']['final_amount']),
    ]
    for col, value in enumerate(totals, start=1):
        cell = ws.cell(total_row, col, value)
        cell.font = Font(bold=True)
        cell.fill = fill_total
        cell.border = thin
        if col in (3, 4, 6, 9) and value != '':
            cell.number_format = '#,##0.00'
        if col == 5 and value != '':
            cell.number_format = '0.0000'

    widths = [6, 32, 12, 18, 14, 18, 16, 16, 20]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Second sheet: English labels matching the paper register
    ws2 = wb.create_sheet('Register style')
    ws2['A1'] = 'Shareholders Profit Distribution'
    ws2['A1'].font = header_font
    ws2.merge_cells('A1:F1')
    ws2['A3'] = 'Faahfaahin / Item'
    ws2['B3'] = 'Lacagta / Amount'
    ws2['A3'].fill = fill_header
    ws2['B3'].fill = fill_header
    ws2['A3'].font = fill_header_font
    ws2['B3'].font = fill_header_font
    ws2['A4'] = 'Faaiidada shirkadda (Company Net Profit)'
    ws2['B4'] = float(data['net_profit'])
    ws2['B4'].number_format = '"$"#,##0.00'
    ws2['A5'] = (
        f'dadka qaar oo lacagta faaiidada '
        f'{100 - float(data["allocation_percent"]):.0f}% kaliya laga qaado '
        f'(Shareholders pool {data["allocation_percent"]}%)'
    )
    ws2['B5'] = float(data['profit_pool'])
    ws2['B5'].number_format = '"$"#,##0.00'

    reg_headers = ['№', 'Magaca', 'Saamiga', 'Lacagta (Capital)', 'Boqoleey %', 'Lacagta (Profit)']
    for col, title in enumerate(reg_headers, start=1):
        cell = ws2.cell(7, col, title)
        cell.fill = fill_header
        cell.font = fill_header_font
        cell.border = thin

    for i, row in enumerate(data['rows']):
        r = 8 + i
        ws2.cell(r, 1, row['no']).border = thin
        ws2.cell(r, 2, row['name']).border = thin
        c3 = ws2.cell(r, 3, float(row['shares']))
        c3.number_format = '#,##0.####'
        c3.border = thin
        c4 = ws2.cell(r, 4, float(row['capital']))
        c4.number_format = '"$"#,##0.00'
        c4.border = thin
        c5 = ws2.cell(r, 5, float(row['ownership_percent']))
        c5.number_format = '0.00'
        c5.border = thin
        c6 = ws2.cell(r, 6, float(row['final_amount']))
        c6.number_format = '"$"#,##0.00'
        c6.border = thin
        c6.font = money_font

    tr = 8 + len(data['rows'])
    ws2.cell(tr, 2, 'Total').font = Font(bold=True)
    for col in range(1, 7):
        ws2.cell(tr, col).fill = fill_total
        ws2.cell(tr, col).border = thin
    ws2.cell(tr, 3, float(data['totals']['shares'])).number_format = '#,##0.####'
    ws2.cell(tr, 4, float(data['totals']['capital'])).number_format = '"$"#,##0.00'
    ws2.cell(tr, 5, float(data['totals']['ownership_percent'])).number_format = '0.00'
    ws2.cell(tr, 6, float(data['totals']['final_amount'])).number_format = '"$"#,##0.00'
    for col, width in enumerate([6, 34, 12, 18, 12, 18], start=1):
        ws2.column_dimensions[get_column_letter(col)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), f'{data["filename_base"]}.xlsx'


def build_period_distribution_csv(period) -> tuple[str, str]:
    data = build_period_distribution_rows(period)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(['Company', data['company_name']])
    writer.writerow(['Period', data['period_label']])
    writer.writerow(['Status', data['status']])
    writer.writerow(['Odoo reference', data['odoo_reference']])
    writer.writerow([])
    writer.writerow(['Summary'])
    writer.writerow(['Company Net Profit', f'{data["net_profit"]:.2f}'])
    writer.writerow(
        [
            f'Shareholders pool ({data["allocation_percent"]}%)',
            f'{data["profit_pool"]:.2f}',
        ]
    )
    writer.writerow([f'{data["partner_name"]} share', f'{data["partner_share"]:.2f}'])
    writer.writerow(['Total distributed', f'{data["totals"]["final_amount"]:.2f}'])
    writer.writerow([])
    writer.writerow([
        'No.',
        'Shareholder Name',
        'Shares',
        'Capital Investment',
        'Ownership %',
        'Original Profit',
        'Arrangement net',
        'Manual adjustment',
        'Profit Allocation',
    ])
    for row in data['rows']:
        writer.writerow([
            row['no'],
            row['name'],
            f'{row["shares"]:.4f}',
            f'{row["capital"]:.2f}',
            f'{row["ownership_percent"]:.4f}',
            f'{row["base_share"]:.2f}',
            f'{row["arrangement_net"]:.2f}',
            f'{row["manual_adjustment"]:.2f}',
            f'{row["final_amount"]:.2f}',
        ])
    writer.writerow([
        '',
        'TOTAL',
        f'{data["totals"]["shares"]:.4f}',
        f'{data["totals"]["capital"]:.2f}',
        f'{data["totals"]["ownership_percent"]:.4f}',
        f'{data["totals"]["base_share"]:.2f}',
        '',
        '',
        f'{data["totals"]["final_amount"]:.2f}',
    ])
    return buf.getvalue(), f'{data["filename_base"]}.csv'


def build_capital_register_xlsx(*, active_only: bool = True) -> tuple[bytes, str]:
    """Export the current capital register (Magaca / Saamiga / Lacagta / Boqoleey)."""
    from apps.models.shareholder import Shareholder
    from apps.services.brand_service import get_brand_settings
    from apps.services.share_value_service import get_share_settings
    from apps.services.shareholder_service import get_ownership_percent
    from datetime import datetime

    query = Shareholder.query
    if active_only:
        query = query.filter_by(is_active=True)
    shareholders = query.order_by(Shareholder.name).all()
    as_of = datetime.utcnow().date()
    brand = get_brand_settings()
    share_settings = get_share_settings()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Capital Register'
    ws['A1'] = f'{brand.get("company_name") or "Akram Sweets"} — Shareholder Capital Register'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')

    headers = ['№', 'Magaca', 'Saamiga', 'Lacagta', 'Boqoleey %', 'Email', 'Phone', 'Active']
    fill_header = PatternFill('solid', fgColor='8A1B24')
    header_font = Font(bold=True, color='FFFFFF')
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(3, col, title)
        cell.fill = fill_header
        cell.font = header_font

    total_shares = Decimal('0')
    total_capital = Decimal('0')
    total_own = Decimal('0')
    for i, sh in enumerate(shareholders, start=1):
        ownership = _pct(get_ownership_percent(sh, as_of))
        shares = Decimal(str(sh.share_count or 0))
        capital = _money(sh.investment_amount)
        total_shares += shares
        total_capital += capital
        total_own += ownership
        r = 3 + i
        ws.cell(r, 1, i)
        ws.cell(r, 2, sh.name)
        ws.cell(r, 3, float(shares)).number_format = '#,##0.####'
        ws.cell(r, 4, float(capital)).number_format = '"$"#,##0.00'
        ws.cell(r, 5, float(ownership)).number_format = '0.0000'
        ws.cell(r, 6, sh.email or '')
        ws.cell(r, 7, sh.phone or '')
        ws.cell(r, 8, 'Yes' if sh.is_active else 'No')

    tr = 4 + len(shareholders)
    ws.cell(tr, 2, 'Total').font = Font(bold=True)
    ws.cell(tr, 3, float(total_shares)).number_format = '#,##0.####'
    ws.cell(tr, 4, float(total_capital)).number_format = '"$"#,##0.00'
    ws.cell(tr, 5, float(total_own)).number_format = '0.0000'
    ws.cell(tr + 2, 1, 'Murabaha / company-owned assets (reporting only)')
    ws.cell(tr + 2, 4, float(_money(share_settings.get('company_owned_assets')))).number_format = (
        '"$"#,##0.00'
    )

    for col, width in enumerate([6, 32, 12, 16, 12, 28, 16, 10], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    suffix = 'active' if active_only else 'all'
    return buf.getvalue(), f'shareholder_capital_register_{suffix}.xlsx'
